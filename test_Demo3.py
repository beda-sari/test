from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
import pytest
import openpyxl
from constants import globalConstants as c


class Test_DemoClass:
    
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 

    def teardown_method(self): 
        self.driver.quit()


    def getData():      
        excel = openpyxl.load_workbook(c.sauceDemo_login_xlsx)
        sheet = excel["Sheet1"] 
        rows = sheet.max_row 
        data = []
        for i in range(2,rows+1):          
            username = sheet.cell(i,1).value    
            password = sheet.cell(i,2).value
            data.append((username,password)) 

        return data

   
    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        if username is not None:
            usernameInput.send_keys(username)
       
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        if password is not None:
            passwordInput.send_keys(password)
       
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_XPATH)
        if username is None and password is None:
            assert errorMessage.text == c.USERNAME_PASSWORD_EMPTY
        elif username is not None and password is None:
            assert errorMessage.text == c.PASSWORD_EMPTY
        elif username is not None and password is not None:
            assert errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH
        
        
        
    def getData():      
        excel = openpyxl.load_workbook(c.sauceDemo_login_xlsx)
        sheet = excel["Sheet2"] 
        rows = sheet.max_row 
        data = []
        for i in range(2,rows+1):          
            username = sheet.cell(i,1).value    
            password = sheet.cell(i,2).value
            data.append((username,password)) 

        return data
    @pytest.mark.parametrize("username,password",getData())
    def test_valid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        if username is not None:
            usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        if password is not None:
            passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
       
        listOfCourses = self.driver.find_elements(By.CLASS_NAME,c.CLASS_NAME)
        assert len(listOfCourses) == 6

        