from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
import pytest
import openpyxl
from constants import globalConstants as c

class Test_Demo2Class:
    

    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.BACK_PACK_ADD)))
        addToCart.click()
        self.driver.execute_script(c.WİNDOW_SCROLL) 
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.T_RED_ADD)))
        addToCart.click()
        self.driver.execute_script(c.WİNDOW_SCROLL2) 
        shoppingCartLink = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CLASS_NAME,c.CART_LİNK)))
        shoppingCartLink.click()

    def teardown_method(self): 
        self.driver.quit()
    
    def getData():      
        excel = openpyxl.load_workbook(c.your_information_xlsx)
        sheet = excel["Sheet1"] 
        rows = sheet.max_row 
        data = []
        for i in range(2,rows+1):          
            firstName = sheet.cell(i,1).value    
            lastName = sheet.cell(i,2).value
            postalCode = sheet.cell(i,3).value
            data.append((firstName,lastName,postalCode)) 

        return data

    
    def test_add_to_cart(self):
        
        listOfCourses = self.driver.find_elements(By.CLASS_NAME,c.C_NAME)
        assert len(listOfCourses) == 2

    def test_remove_from_cart(self):
        remove = WebDriverWait(self.driver,8).until(ec.visibility_of_element_located((By.XPATH,c.BACK_PACK_REMOVE)))
        remove.click()
        remove = WebDriverWait(self.driver,8).until(ec.visibility_of_element_located((By.XPATH,c.T_RED_REMOVE)))
        remove.click()
        listOfCourses = self.driver.find_elements(By.CLASS_NAME,c.C_NAME)
        assert len(listOfCourses) == 0

    @pytest.mark.parametrize("firstName,lastName,postalCode",getData())
    
    def test_checkout(self,firstName,lastName,postalCode):
        checkout = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.CHECKOUT_ID)))
        checkout.click()
        firstNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.FIRST_NAME_ID)))
        firstNameInput.send_keys(firstName)
       
        lastNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.LAST_NAME_ID)))
        lastNameInput.send_keys(lastName)

        postalCodeInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.POSTAL_CODE_ID)))
        
        if postalCode is not None:
            postalCodeInput.send_keys(postalCode)

        continueButton = self.driver.find_element(By.ID,c.CONTINUE_ID)
        continueButton.click()
       
        errorMessage = self.driver.find_element(By.XPATH,c.CONTINUE_XPATH)
        assert errorMessage.text == c.ERROR_POSTAL_CODE
    